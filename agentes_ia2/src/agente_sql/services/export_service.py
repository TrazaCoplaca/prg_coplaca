from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle


def _normalizar_valor(valor: Any) -> str:
    if valor is None:
        return ""
    return str(valor)


def generar_excel_bytes(
    columnas: list[str],
    filas: list[dict[str, Any]],
    nombre_hoja: str = "Resultados",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja

    # Cabecera
    ws.append(columnas)

    # Filas
    for fila in filas:
        ws.append([fila.get(col) for col in columnas])

    # Ajuste simple de ancho
    for i, columna in enumerate(columnas, start=1):
        max_len = len(columna)
        for fila in filas:
            valor = _normalizar_valor(fila.get(columna))
            if len(valor) > max_len:
                max_len = len(valor)

        ancho = min(max_len + 2, 40)
        ws.column_dimensions[get_column_letter(i)].width = ancho

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generar_pdf_bytes(
    columnas: list[str],
    filas: list[dict[str, Any]],
) -> bytes:
    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    data = [columnas]
    for fila in filas:
        data.append([_normalizar_valor(fila.get(col)) for col in columnas])

    tabla = Table(data, repeatRows=1)

    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9ead3")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
            ]
        )
    )

    doc.build([tabla])
    output.seek(0)
    return output.getvalue()