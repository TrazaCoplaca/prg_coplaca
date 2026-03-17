"""
Rutas de exportación de datos
"""
import io
from typing import List, Dict, Any, Set
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from agente_sql.infrastructure.logger import get_logger

logger = get_logger(__name__)

router = APIRouter(prefix="/export")


class ExportRequest(BaseModel):
    """Modelo para solicitudes de exportación"""
    data: List[Dict[str, Any]]
    filename: str = "export"
    title: str = ""


# Columnas a EXCLUIR del PDF (mantener en Excel)
PDF_EXCLUDED_COLUMNS: Set[str] = {
    # IDs internos
    'IdTrazaCaja', 'IdCaja', 'IdAlbaran', 'IdSocio', 'IdVariedad', 
    'IdCategoria', 'IdCalidad', 'IdProducto', 'Id',
    
    # Campos técnicos
    'Traspasado', 'FechaModificacion', 'UsuarioModificacion', 
    'FechaCreacion', 'UsuarioCreacion', 'Activo', 'Eliminado',
    'Version', 'Sincronizado', 'Hash', 'Timestamp', 'RowVersion',
    
    # Otros campos no relevantes para trazabilidad
    'Interno', 'Sistema', 'Migrado', 'IdExterno'
}

# Columnas prioritarias para PDF (orden de importancia)
PDF_PRIORITY_COLUMNS = [
    'CodigoCaja', 'NumeroCaja', 'Caja',
    'NumeroAlbaran', 'CodigoAlbaran', 'Albaran',
    'CodigoSocio', 'NombreSocio', 'Socio',
    'FechaAlbaran', 'Fecha',
    'Variedad', 'TipoProducto', 'Producto',
    'Peso', 'Kilos', 'PesoBruto', 'PesoNeto',
    'Calidad', 'Categoria', 'Clasificacion',
    'Estado', 'EstadoCaja',
    'Lote', 'NumeroLote',
    'Origen', 'Procedencia',
    'Destino',
    'Observaciones', 'Notas'
]


def filter_columns_for_pdf(columns: List[str], max_columns: int = 8) -> List[str]:
    """
    Filtra y ordena las columnas para el PDF según prioridad.
    
    Args:
        columns: Lista de todas las columnas disponibles
        max_columns: Número máximo de columnas a incluir en PDF
        
    Returns:
        Lista filtrada y ordenada de columnas
    """
    # Excluir columnas no deseadas
    filtered = [col for col in columns if col not in PDF_EXCLUDED_COLUMNS]
    
    # Ordenar según prioridad
    priority_filtered = []
    remaining = []
    
    for col in filtered:
        if col in PDF_PRIORITY_COLUMNS:
            priority_filtered.append((PDF_PRIORITY_COLUMNS.index(col), col))
        else:
            remaining.append(col)
    
    # Ordenar por prioridad
    priority_filtered.sort(key=lambda x: x[0])
    priority_columns = [col for _, col in priority_filtered]
    
    # Combinar: primero las prioritarias, luego las restantes
    final_columns = priority_columns + remaining
    
    # Limitar al número máximo
    return final_columns[:max_columns]


@router.post("/excel", summary="Exportar a Excel")
async def export_excel(request: ExportRequest):
    """
    Exporta los datos proporcionados a formato Excel (.xlsx)
    Incluye TODAS las columnas
    """
    try:
        if not request.data:
            raise HTTPException(status_code=400, detail="No hay datos para exportar")

        logger.info(f"Iniciando exportación Excel con {len(request.data)} filas")
        logger.debug(f"Primera fila de datos: {request.data[0]}")

        # Crear un archivo Excel en memoria
        output = io.BytesIO()
        
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Trazabilidad"

        # Obtener TODAS las columnas
        if isinstance(request.data[0], dict):
            columns = list(request.data[0].keys())
        else:
            raise HTTPException(status_code=400, detail="Formato de datos no válido")

        logger.info(f"Columnas Excel: {len(columns)} columnas - {columns}")

        # Estilos para el encabezado
        header_fill = PatternFill(start_color="00964A", end_color="00964A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin', color='D8E2DA'),
            right=Side(style='thin', color='D8E2DA'),
            top=Side(style='thin', color='D8E2DA'),
            bottom=Side(style='thin', color='D8E2DA')
        )

        # Estilos para celdas de datos
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        data_font = Font(size=10)

        # Escribir encabezados
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = str(column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Escribir datos
        for row_idx, row_data in enumerate(request.data, start=2):
            for col_idx, column in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Obtener el valor
                value = row_data.get(column)
                
                # Manejar diferentes tipos de datos
                if value is None or value == "":
                    cell.value = ""
                elif isinstance(value, (int, float)):
                    cell.value = value
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                elif isinstance(value, bool):
                    cell.value = "Sí" if value else "No"
                else:
                    cell.value = str(value).strip()
                
                cell.font = data_font
                cell.border = thin_border
                
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F4F1", end_color="F0F4F1", fill_type="solid")

        # Ajustar ancho de columnas
        for col_idx, column in enumerate(columns, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(column))
            
            for row_idx in range(2, min(len(request.data) + 2, 100)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_length = len(str(cell_value))
                    max_length = max(max_length, cell_length)
            
            adjusted_width = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Congelar primera fila
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Guardar
        wb.save(output)
        output.seek(0)

        filename = f"{request.filename}.xlsx"
        logger.info(f"Excel generado: {filename} ({len(request.data)} filas, {len(columns)} columnas)")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Error al exportar a Excel: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")


@router.post("/pdf", summary="Exportar a PDF")
async def export_pdf(request: ExportRequest):
    """
    Exporta los datos proporcionados a formato PDF
    Solo incluye columnas RELEVANTES para trazabilidad
    """
    try:
        if not request.data:
            raise HTTPException(status_code=400, detail="No hay datos para exportar")

        logger.info(f"Iniciando exportación PDF con {len(request.data)} filas")

        # Crear un archivo PDF en memoria
        output = io.BytesIO()
        
        # Obtener todas las columnas
        all_columns = list(request.data[0].keys())
        
        # Filtrar columnas para PDF (máximo 8 columnas importantes)
        columns = filter_columns_for_pdf(all_columns, max_columns=8)
        
        logger.info(f"Columnas PDF filtradas ({len(columns)}): {columns}")
        logger.info(f"Columnas excluidas: {set(all_columns) - set(columns)}")
        
        # Configurar documento siempre en horizontal para tablas
        pagesize = landscape(A4)
        
        doc = SimpleDocTemplate(
            output,
            pagesize=pagesize,
            rightMargin=15,
            leftMargin=15,
            topMargin=30,
            bottomMargin=18,
        )

        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#00964A'),
            spaceAfter=12,
            alignment=1
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#5e6b5e'),
            spaceAfter=10,
            alignment=1
        )

        # Título
        if request.title:
            title = Paragraph(request.title, title_style)
            elements.append(title)
        else:
            title = Paragraph("Trazabilidad de Cajas", title_style)
            elements.append(title)
        
        # Subtítulo
        subtitle_text = f"Total de registros: {len(request.data)} | Columnas mostradas: {len(columns)}"
        subtitle = Paragraph(subtitle_text, subtitle_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.15 * inch))

        # Preparar datos para la tabla (SOLO columnas filtradas)
        table_data = [columns]  # Encabezados
        
        for row in request.data:
            table_row = []
            for col in columns:
                value = row.get(col)
                
                # Formatear valores
                if value is None or value == "":
                    str_value = "-"
                elif isinstance(value, bool):
                    str_value = "Sí" if value else "No"
                elif isinstance(value, (int, float)):
                    str_value = str(value)
                else:
                    str_value = str(value).strip()
                
                # Limitar longitud
                if len(str_value) > 35:
                    str_value = str_value[:32] + "..."
                
                table_row.append(str_value)
            table_data.append(table_row)

        # Calcular anchos de columnas proporcionalmente
        available_width = pagesize[0] - 30  # Restar márgenes
        col_width = available_width / len(columns)
        col_widths = [col_width] * len(columns)

        # Crear tabla
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Estilos de la tabla
        table_style = TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00964A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F1')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D8E2DA')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ])

        table.setStyle(table_style)
        elements.append(table)

        # Pie de página
        from datetime import datetime
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.HexColor('#7b877c'),
            alignment=2
        )
        elements.append(Spacer(1, 0.2 * inch))
        footer_text = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} | Para ver todos los campos, exportar a Excel"
        footer = Paragraph(footer_text, footer_style)
        elements.append(footer)

        # Generar PDF
        doc.build(elements)
        output.seek(0)

        filename = f"{request.filename}.pdf"
        logger.info(f"PDF generado: {filename} ({len(request.data)} filas, {len(columns)} columnas)")

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Error al exportar a PDF: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")


@router.post("/csv", summary="Exportar a CSV")
async def export_csv(request: ExportRequest):
    """
    Exporta los datos proporcionados a formato CSV
    Incluye TODAS las columnas
    """
    try:
        if not request.data:
            raise HTTPException(status_code=400, detail="No hay datos para exportar")

        logger.info(f"Iniciando exportación CSV con {len(request.data)} filas")

        # Obtener TODAS las columnas
        columns = list(request.data[0].keys())
        
        # Crear CSV
        output = io.StringIO()
        
        # Escribir encabezados
        output.write(','.join(f'"{col}"' for col in columns) + '\n')
        
        # Escribir datos
        for row in request.data:
            values = []
            for col in columns:
                value = row.get(col)
                
                if value is None or value == "":
                    values.append('""')
                elif isinstance(value, bool):
                    values.append('"Sí"' if value else '"No"')
                elif isinstance(value, (int, float)):
                    values.append(str(value))
                else:
                    str_value = str(value).replace('"', '""')
                    values.append(f'"{str_value}"')
            
            output.write(','.join(values) + '\n')
        
        output.seek(0)

        # Convertir a BytesIO con BOM
        bytes_output = io.BytesIO()
        bytes_output.write('\ufeff'.encode('utf-8'))
        bytes_output.write(output.getvalue().encode('utf-8'))
        bytes_output.seek(0)

        filename = f"{request.filename}.csv"
        logger.info(f"CSV generado: {filename} ({len(request.data)} filas, {len(columns)} columnas)")

        return StreamingResponse(
            bytes_output,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Error al exportar a CSV: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al generar CSV: {str(e)}")